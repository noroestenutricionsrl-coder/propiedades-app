import os
import json
import tempfile
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def estilo_header(ws, fila, columnas, color="1a5c3a"):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    for col in range(1, columnas + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def borde_fino():
    lado = Side(style='thin', color='CCCCCC')
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def exportar_vencimientos_excel(mes=None, anio=None):
    """Genera un Excel con vencimientos del mes/año especificado."""
    from .models import Vencimiento
    
    hoy = date.today()
    mes = mes or hoy.month
    anio = anio or hoy.year
    
    vencimientos = Vencimiento.objects.filter(
        fecha_vencimiento__year=anio,
        fecha_vencimiento__month=mes
    ).select_related(
        'propiedad_servicio__propiedad',
        'propiedad_servicio__servicio'
    ).order_by('fecha_vencimiento', 'propiedad_servicio__propiedad__domicilio')
    
    wb = Workbook()
    ws = wb.active
    
    nombre_mes = date(anio, mes, 1).strftime('%B %Y').capitalize()
    ws.title = nombre_mes
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f'Vencimientos — {nombre_mes}'
    ws['A1'].font = Font(size=14, bold=True, color='1a5c3a')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Generado el {hoy.strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(size=9, color='888888', italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Propiedad', 'Servicio/Impuesto', 'Tipo', 'Vencimiento', 'Importe', 'Responsable', 'Estado']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    estilo_header(ws, 4, len(headers))
    ws.row_dimensions[4].height = 22
    
    # Datos
    colores = {
        'pagado': 'E8F5EE',
        'vencido': 'FDECEA',
        'pendiente': 'FEF9E7',
    }
    
    for row_idx, v in enumerate(vencimientos, 5):
        ws.cell(row=row_idx, column=1, value=v.propiedad_servicio.propiedad.domicilio)
        ws.cell(row=row_idx, column=2, value=v.propiedad_servicio.servicio.nombre)
        ws.cell(row=row_idx, column=3, value=v.propiedad_servicio.servicio.get_tipo_display())
        ws.cell(row=row_idx, column=4, value=v.fecha_vencimiento.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=5, value=float(v.importe_estimado) if v.importe_estimado else '')
        ws.cell(row=row_idx, column=6, value=v.propiedad_servicio.get_responsable_pago_display())
        ws.cell(row=row_idx, column=7, value=v.get_estado_display())
        
        color = colores.get(v.estado, 'FFFFFF')
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.border = borde_fino()
            cell.alignment = Alignment(vertical='center')
    
    # Formato importe
    for row in range(5, 5 + vencimientos.count()):
        ws.cell(row=row, column=5).number_format = '$ #,##0.00'
    
    # Ancho columnas
    anchos = [40, 30, 12, 15, 18, 15, 12]
    for col, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    
    # Guardar en archivo temporal
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name, f'Vencimientos_{nombre_mes.replace(" ", "_")}.xlsx'


def subir_a_drive(ruta_archivo, nombre_archivo):
    """Sube un archivo a Google Drive."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    
    folder_id = os.environ.get('GOOGLE_DRIVE_FOLDER_ID', '')
    credenciales_json = os.environ.get('GOOGLE_CREDENTIALS_JSON', '')
    
    if not credenciales_json or not folder_id:
        return None, "Credenciales de Google Drive no configuradas"
    
    try:
        creds_dict = json.loads(credenciales_json)
        creds = service_account.Credentials.from_service_account_info(
            creds_dict,
            scopes=['https://www.googleapis.com/auth/drive.file']
        )
        service = build('drive', 'v3', credentials=creds)
        
        file_metadata = {
            'name': nombre_archivo,
            'parents': [folder_id]
        }
        media = MediaFileUpload(
            ruta_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,webViewLink'
        ).execute()
        
        return file.get('webViewLink'), None
    except Exception as e:
        return None, str(e)
    finally:
        if os.path.exists(ruta_archivo):
            os.unlink(ruta_archivo)
